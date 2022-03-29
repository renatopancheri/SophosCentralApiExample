#!/usr/bin/env python3

import json
import urllib3
from urllib.parse import urlencode
import openpyxl
from pathlib import Path
import time

# endpoint for token JWT used to populate Authorization header in all successive http requests
jwt_endpoint = 'https://id.sophos.com/api/v2/oauth2/token'
# used to retrieve the organization uuid
organizations_endpoint = 'https://api.central.sophos.com/whoami/v1'
# used to retrieve the Company tenant
tenants_endpoint = 'https://api.central.sophos.com/organization/v1/tenants'
endpoints_endpoint = '/endpoint/v1/endpoints'
tamper_get_endpoint = '/endpoint/v1/endpoints/{endpointId}/tamper-protection'
users_endpoint = '/common/v1/directory/users'
tenant_name = 'My Company'
excel_file = 'MyExcel.xlsx'
excel_groups_to_match = ['group A', 'group B']

http = urllib3.PoolManager()

def api_request(request_type, endpoint, headers, body=None):
    if body is None:
        r = http.request(request_type, endpoint, headers=headers)
    else:
        r = http.request(request_type, endpoint, headers=headers,body=body)
    return json.loads(r.data)

# initialization api requests to get jwt token, org id and tenant id
def populate_sophos_data(client_id, secret_id, tenant_name):
    sophos_data = {}
    sophos_data['client_id'] = client_id
    sophos_data['client_secret'] = secret_id

    # JWT TOKEN
    jwt_body = {"grant_type": "client_credentials",
        "client_id": sophos_data['client_id'], "client_secret": sophos_data['client_secret'],
        "scope": "token"}
    encoded_body = urlencode(jwt_body)
    r = api_request('POST', jwt_endpoint, {'Content-Type': 'application/x-www-form-urlencoded'}, encoded_body)
    sophos_data['access_token'] = 'Bearer ' + r['access_token']
    print('JWT Token: ' + sophos_data['access_token'])
    # ORGANIZATION UUID
    r = api_request('GET', organizations_endpoint, {'Authorization': sophos_data['access_token']})
    sophos_data['organization_uuid'] = r["id"]
    print('Organization UUID: ' +sophos_data['organization_uuid'])
    # COMPANY TENANT UUID
    r = api_request('GET', tenants_endpoint,
        {'X-Organization-ID': sophos_data['organization_uuid'], 'Authorization': sophos_data['access_token']})
    for elem in r['items']:
        if elem['name'] == tenant_name:
            company_endpoint = elem
    sophos_data['regional_endpoint'] = company_endpoint['apiHost']
    sophos_data['company_tenant_uuid'] = company_endpoint['id']
    print('Tenant UUID: ' +  sophos_data['company_tenant_uuid'])
    return sophos_data
    
# get endpoints matching query parameter
def get_endpoints(sophos_data, query=None):
    base_endpoint = sophos_data['regional_endpoint'] + endpoints_endpoint
    if query is None:
        query = {}
#    if query is not None:
    endpoint = base_endpoint + '?' + urlencode(query)
    print('searching for ' + urlencode(query))
    print(endpoint)
    ret = []
    r = api_request('GET', endpoint,
        {'X-Tenant-ID': sophos_data['company_tenant_uuid'],
        'Authorization':  sophos_data['access_token']}
    )
    ret = ret + r['items']
    new_page = 'nextKey' in r['pages']
    while new_page:
        query['pageFromKey'] = r['pages']['nextKey']
        endpoint = base_endpoint + '?' + urlencode(query)
        r = api_request('GET', endpoint,
            {'X-Tenant-ID': sophos_data['company_tenant_uuid'],
            'Authorization':  sophos_data['access_token']}
        )
        new_page = 'nextKey' in r['pages']
        ret = ret + r['items']
        time.sleep(0.2)
#    print('Found ' + str(len(r['items'])) + ' items')
    return ret

# get users matching query parameter
def get_users(sophos_data, query=None):
    endpoint = sophos_data['regional_endpoint'] + users_endpoint
    if query is not None:
        endpoint = endpoint + '?' + urlencode(query)
        print('searching for ' + urlencode(query))
    print(endpoint)
    r = api_request('GET', endpoint,
        {'X-Tenant-ID': sophos_data['company_tenant_uuid'],
        'Authorization':  sophos_data['access_token']}
    )
#    print('Found ' + str(len(r['items'])) + ' items')
    return r['items']

# get all endpoints connected to an exchange login
def get_endpoints_from_user(sophos_data, username):
    user = []
    r = get_users(sophos_data, {'search': username, 'searchFields': 'exchangeLogin'})
    if len(r) == 1:
        user = r[0]
        print("Found 1 User: " + user['exchangeLogin'] + "    " + user['id'] + "    " + user['name'])
        #r = get_endpoints(sophos_data, {"associatedPersonContains": user['name']})
        temp = get_endpoints(sophos_data, {"search": user['name'], 'searchFields': 'associatedPersonName'})
        r = []
        for elem in temp:
            if username in elem['associatedPerson']['viaLogin']:
                r.append(elem)
        temp = get_endpoints(sophos_data, {"search": user['exchangeLogin'], 'searchFields': 'associatedPersonName'})
        for elem in temp:
            if username in elem['associatedPerson']['viaLogin']:
                r.append(elem)
        print("Found " + str(len(r)) + " endpoints for User " + username + ":")
    else:
        print("Something went wrong when searching user " + username)
        r = []
    return r

# get tamper status from endpoint id
def get_tamper(sophos_data, endpoint_id):
    r = api_request('GET', sophos_data['regional_endpoint'] + 
        tamper_get_endpoint.replace('{endpointId}', endpoint_id),
        {'X-Tenant-ID': sophos_data['company_tenant_uuid'],
        'Authorization': sophos_data['access_token']})
    return r

def set_tamper(sophos_data, endpoint_id, enabled):
    r = api_request('POST', sophos_data['regional_endpoint'] +
        tamper_get_endpoint.replace('{endpointId}', endpoint_id),
        {'X-Tenant-ID': sophos_data['company_tenant_uuid'],
        'Authorization': sophos_data['access_token']},
        json.dumps({"enabled": enabled, "regeneratePassword": False}))
    print(r)

    #5a5ebd51-e4d1-4f3b-b03a-a6476e6fe3f6

# get tamper of all endpoints matching an exchange login
def get_tamper_from_username(sophos_data, username):
    print("")
    print("")
    r = get_endpoints_from_user(sophos_data, username)
    for elem in r:
        tamper = get_tamper(sophos_data, elem['id'])
        print(elem['hostname']+ "   " + elem['id'] + "   " + str(elem['associatedPerson']) + str(tamper))

#set_tamper_from_username(sophos_data, username, status):
#    r = get_endpoints_from_user(sophos_data, username)
#    for elem in r:
#        tamper =

# read credentials from credentials.json file and do the first api request
# to get token jwt, organization id and company tenant id
sophos_data = {}
with open('credentials.json', 'r') as credentials:
    credentials_data = json.load(credentials)
    sophos_data = populate_sophos_data(
        credentials_data["CLIENT_ID"], credentials_data["CLIENT_SECRET"],
        tenant_name)

xlsx_file = Path(excel_file)
wb = openpyxl.load_workbook(xlsx_file)
# open second sheet
ws = wb.worksheets[1]
all_endpoints = get_endpoints(sophos_data, {'pageSize': 500})
for row in range(2,ws.max_row+1):
    cell_name = "D" + str(row)
    if ws[cell_name].value in excel_groups_to_match:
        cell_name = "B" + str(row)
        print(ws[cell_name].value)
        endpoints = []
        for elem in all_endpoints:
            if 'associatedPerson' in elem:
                if 'name' in elem['associatedPerson'] and ws[cell_name].value in elem['associatedPerson']['name']:
                    endpoints.append(elem)
                elif 'viaLogin' in elem['associatedPerson'] and ws[cell_name].value in elem['associatedPerson']['viaLogin']:
                    endpoints.append(elem)
        #print(endpoints)
        for elem in endpoints:
            print("   " + elem['hostname'] + ":   "+ str(get_tamper(sophos_data, elem['id'])))
            time.sleep(0.1)
            res = set_tamper(sophos_data, elem['id'], False)
            print("   " + elem['hostname'] + ": Tamper has been Disabled")
            time.sleep(0.1)
#    get_tamper_from_username(sophos_data, (ws[cell_name].value))
#get_tamper_from_username(sophos_data, 'u09689')
#set_tamper(sophos_data, '5a5ebd51-e4d1-4f3b-b03a-a6476e6fe3f6', False)
#set_tamper_from_username(sophos_data, 'u09689', True)

